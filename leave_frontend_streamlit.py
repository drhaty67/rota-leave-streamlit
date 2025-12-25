import streamlit as st
from datetime import date
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd

st.set_page_config(page_title="Rota Leave Entry", layout="centered")

TEMPLATE_DEFAULT = "Rota_Publish_Template_ORtools.xlsx"

st.title("Rota Leave Entry")
st.write("Enter leave requests and write them into the **Leave** sheet of the rota template workbook.")

# --- File selection ---
st.subheader("1) Select workbook")
uploaded = st.file_uploader("Upload your Rota_Publish_Template_ORtools.xlsx", type=["xlsx"])

if uploaded is None:
    st.info(f"Upload your workbook (default filename: {TEMPLATE_DEFAULT}).")
    st.stop()

# Save uploaded workbook to a local temp file
tmp_path = Path("uploaded_rota.xlsx")
tmp_path.write_bytes(uploaded.getbuffer())

wb = load_workbook(tmp_path)
if "Leave" not in wb.sheetnames or "Consultants" not in wb.sheetnames:
    st.error("Workbook must contain sheets named 'Leave' and 'Consultants'.")
    st.stop()

# --- Read consultants (active only) ---
cws = wb["Consultants"]
names = []
for r in range(2, 2000):
    nm = cws[f"A{r}"].value
    active = cws[f"F{r}"].value
    if nm and bool(active):
        names.append(str(nm))
names = sorted(names)

# --- Read existing leave into a table ---
lws = wb["Leave"]
existing = []
for r in range(2, 5000):
    nm = lws[f"A{r}"].value
    if nm is None or nm == "":
        continue
    existing.append({
        "Name": str(nm),
        "StartDate": lws[f"B{r}"].value,
        "EndDate": lws[f"C{r}"].value,
        "LeaveType": lws[f"D{r}"].value,
        "Approved": bool(lws[f"E{r}"].value) if lws[f"E{r}"].value is not None else False,
    })

st.subheader("2) Add leave request")
with st.form("leave_form"):
    col1, col2 = st.columns(2)
    with col1:
        name = st.selectbox("Consultant", options=names)
        leave_type = st.selectbox("Leave type", options=["Annual", "Study", "NOC"])
    with col2:
        start_date = st.date_input("Date from", value=date.today())
        end_date = st.date_input("Date to", value=date.today())
    approved = st.checkbox("Approved", value=True)
    submitted = st.form_submit_button("Add to workbook")

def find_next_empty_row(ws, start_row=2, col="A", max_row=20000):
    for rr in range(start_row, max_row + 1):
        if ws[f"{col}{rr}"].value in (None, ""):
            return rr
    return max_row + 1

if submitted:
    if end_date < start_date:
        st.error("Date to cannot be earlier than Date from.")
    else:
        rr = find_next_empty_row(lws, start_row=2, col="A")
        lws[f"A{rr}"].value = name
        lws[f"B{rr}"].value = start_date
        lws[f"C{rr}"].value = end_date
        lws[f"D{rr}"].value = leave_type
        lws[f"E{rr}"].value = bool(approved)
        wb.save(tmp_path)
        st.success(f"Added leave request for {name}: {start_date} → {end_date} ({leave_type}).")
        st.rerun()

st.subheader("3) Current leave entries")
df = pd.DataFrame(existing)
if df.empty:
    st.write("No leave entries found.")
else:
    st.dataframe(df, width=True, hide_index=True)

st.subheader("4) Download updated workbook")
with open(tmp_path, "rb") as f:
    st.download_button(
        label="Download updated rota workbook",
        data=f,
        file_name="Rota_Publish_Template_ORtools_UPDATED.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
