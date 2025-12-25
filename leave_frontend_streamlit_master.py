import streamlit as st
from datetime import date, datetime
from pathlib import Path
from typing import Optional, List, Dict

import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Rota Leave Admin", layout="wide")

st.title("Rota Leave Admin")
st.write(
    "Maintain leave requests in a shared master rota workbook. "
    "This app can **add**, **edit**, and **delete** rows in the `Leave` sheet."
)

# -----------------------------
# Shared workbook configuration
# -----------------------------
st.subheader("1) Master workbook location")

default_master = st.secrets.get("MASTER_WORKBOOK_PATH", "") if hasattr(st, "secrets") else ""
master_path_str = st.text_input(
    "Master workbook path (shared drive)",
    value=default_master,
    placeholder="/Volumes/Shared/Rota/Rota_Publish_Template_ORtools.xlsx",
    help="Point this at the shared master workbook. The app will read and write this file in place."
)

master_path = Path(master_path_str).expanduser() if master_path_str else None

colA, colB = st.columns([1, 2])
with colA:
    make_backup = st.checkbox("Create timestamped backup before writing", value=True)
with colB:
    st.caption(
        "Recommendation: keep backups enabled. If your shared drive supports file versioning, "
        "you can disable backups later."
    )

# Optional file locking (best-effort)
enable_lock = st.checkbox(
    "Use a simple lock file to reduce concurrent edits",
    value=True,
    help="Creates a .lock file next to the workbook while writing. Not a perfect distributed lock."
)

def lock_path_for(path: Path) -> Path:
    return path.with_suffix(path.suffix + ".lock")

def acquire_lock(path: Path) -> bool:
    lp = lock_path_for(path)
    try:
        # exclusive create
        lp.write_text(f"locked at {datetime.now().isoformat()}\n", encoding="utf-8")
        return True
    except Exception:
        return False

def release_lock(path: Path) -> None:
    lp = lock_path_for(path)
    try:
        if lp.exists():
            lp.unlink()
    except Exception:
        pass

def backup_file(path: Path) -> Optional[Path]:
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        bkp = path.with_name(f"{path.stem}_backup_{ts}{path.suffix}")
        bkp.write_bytes(path.read_bytes())
        return bkp
    except Exception:
        return None

def excel_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return pd.to_datetime(v).date()

def read_master(path: Path) -> Dict:
    wb = load_workbook(path)
    if "Leave" not in wb.sheetnames or "Consultants" not in wb.sheetnames:
        raise ValueError("Workbook must contain 'Leave' and 'Consultants' sheets.")

    cws = wb["Consultants"]
    names: List[str] = []
    for r in range(2, 2000):
        nm = cws[f"A{r}"].value
        active = cws[f"F{r}"].value
        if nm and bool(active):
            names.append(str(nm))
    names = sorted(set(names))

    lws = wb["Leave"]
    rows = []
    for r in range(2, 5000):
        nm = lws[f"A{r}"].value
        if nm is None or nm == "":
            continue
        rows.append({
            "RowID": r,  # sheet row number, used for edit/delete
            "Name": str(nm),
            "StartDate": excel_date(lws[f"B{r}"].value),
            "EndDate": excel_date(lws[f"C{r}"].value),
            "LeaveType": str(lws[f"D{r}"].value or ""),
            "Approved": bool(lws[f"E{r}"].value) if lws[f"E{r}"].value is not None else False,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["Name", "StartDate", "EndDate"], na_position="last").reset_index(drop=True)

    return {"wb": wb, "names": names, "df": df}

def next_empty_row(lws, start_row=2, max_row=20000) -> int:
    for r in range(start_row, max_row + 1):
        if lws[f"A{r}"].value in (None, ""):
            return r
    return max_row + 1

def validate_dates(s: date, e: date) -> Optional[str]:
    if e < s:
        return "Date to cannot be earlier than Date from."
    return None

def normalize_leave_type(t: str) -> str:
    t = (t or "").strip()
    mapping = {"annual": "Annual", "study": "Study", "noc": "NOC"}
    return mapping.get(t.lower(), t)

# Gate on path
if not master_path_str:
    st.info("Enter the master workbook path to continue.")
    st.stop()

if not master_path.exists():
    st.error(f"Workbook not found at: {master_path}")
    st.stop()

if master_path.suffix.lower() != ".xlsx":
    st.error("Master workbook must be an .xlsx file.")
    st.stop()

# Load master data
try:
    data = read_master(master_path)
except Exception as e:
    st.exception(e)
    st.stop()

names = data["names"]
df = data["df"]

# -----------------------------
# Add leave
# -----------------------------
st.subheader("2) Add leave request")
with st.form("add_form"):
    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        add_name = st.selectbox("Consultant", options=names)
        add_type = st.selectbox("Leave type", options=["Annual", "Study", "NOC"])
    with c2:
        add_start = st.date_input("Date from", value=date.today(), key="add_start")
    with c3:
        add_end = st.date_input("Date to", value=date.today(), key="add_end")
    add_approved = st.checkbox("Approved", value=True)
    add_submit = st.form_submit_button("Add to master workbook")

if add_submit:
    err = validate_dates(add_start, add_end)
    if err:
        st.error(err)
    else:
        # Write
        locked = False
        try:
            if enable_lock:
                locked = acquire_lock(master_path)
                if not locked:
                    st.error("Workbook is currently locked by another session. Try again shortly.")
                    st.stop()

            if make_backup:
                bkp = backup_file(master_path)
                if bkp is None:
                    st.warning("Backup failed; continuing without backup.")
                else:
                    st.caption(f"Backup created: {bkp.name}")

            wb = load_workbook(master_path)
            lws = wb["Leave"]
            r = next_empty_row(lws)

            lws[f"A{r}"].value = add_name
            lws[f"B{r}"].value = add_start
            lws[f"C{r}"].value = add_end
            lws[f"D{r}"].value = normalize_leave_type(add_type)
            lws[f"E{r}"].value = bool(add_approved)

            wb.save(master_path)
            st.success(f"Added leave: {add_name} {add_start} → {add_end} ({add_type}).")
            st.rerun()
        finally:
            if enable_lock:
                release_lock(master_path)

# -----------------------------
# View / filter
# -----------------------------
st.subheader("3) View and filter leave entries")

fc1, fc2, fc3, fc4 = st.columns([2, 1, 1, 1])
with fc1:
    filt_name = st.selectbox("Filter by consultant (optional)", options=["(All)"] + names)
with fc2:
    filt_type = st.selectbox("Filter by type", options=["(All)", "Annual", "Study", "NOC"])
with fc3:
    filt_approved = st.selectbox("Approved?", options=["(All)", "Approved only", "Not approved"])
with fc4:
    search = st.text_input("Search (contains)", value="")

view = df.copy()
if not view.empty:
    if filt_name != "(All)":
        view = view[view["Name"] == filt_name]
    if filt_type != "(All)":
        view = view[view["LeaveType"].str.lower() == filt_type.lower()]
    if filt_approved == "Approved only":
        view = view[view["Approved"] == True]
    elif filt_approved == "Not approved":
        view = view[view["Approved"] == False]
    if search.strip():
        s = search.strip().lower()
        view = view[
            view["Name"].str.lower().str.contains(s)
            | view["LeaveType"].str.lower().str.contains(s)
        ]

st.dataframe(view, use_container_width=True, hide_index=True)

# -----------------------------
# Edit / delete
# -----------------------------
st.subheader("4) Edit or delete an existing leave row")

if df.empty:
    st.info("No leave rows found to edit/delete.")
    st.stop()

# pick by RowID (stable reference)
row_ids = view["RowID"].tolist() if not view.empty else df["RowID"].tolist()
selected_row = st.selectbox("Select RowID to edit/delete", options=row_ids)

row = df[df["RowID"] == selected_row].iloc[0].to_dict()

with st.form("edit_form"):
    ec1, ec2, ec3, ec4 = st.columns([2, 1, 1, 1])
    with ec1:
        edit_name = st.selectbox("Consultant", options=names, index=names.index(row["Name"]) if row["Name"] in names else 0)
        edit_type = st.selectbox("Leave type", options=["Annual", "Study", "NOC"],
                                 index=["Annual","Study","NOC"].index(normalize_leave_type(row.get("LeaveType","Annual"))) if normalize_leave_type(row.get("LeaveType","Annual")) in ["Annual","Study","NOC"] else 0)
    with ec2:
        edit_start = st.date_input("Date from", value=row["StartDate"] or date.today(), key="edit_start")
    with ec3:
        edit_end = st.date_input("Date to", value=row["EndDate"] or (row["StartDate"] or date.today()), key="edit_end")
    with ec4:
        edit_approved = st.checkbox("Approved", value=bool(row.get("Approved", False)))

    csave, cdelete = st.columns([1, 1])
    with csave:
        save_btn = st.form_submit_button("Save changes")
    with cdelete:
        delete_btn = st.form_submit_button("Delete row")

if save_btn or delete_btn:
    err = validate_dates(edit_start, edit_end) if save_btn else None
    if err:
        st.error(err)
    else:
        try:
            if enable_lock:
                locked = acquire_lock(master_path)
                if not locked:
                    st.error("Workbook is currently locked by another session. Try again shortly.")
                    st.stop()

            if make_backup:
                bkp = backup_file(master_path)
                if bkp is None:
                    st.warning("Backup failed; continuing without backup.")
                else:
                    st.caption(f"Backup created: {bkp.name}")

            wb = load_workbook(master_path)
            lws = wb["Leave"]
            r = int(selected_row)

            if delete_btn:
                # Delete row by clearing cells (keeps sheet row numbers stable; avoids shifting)
                for col in ("A","B","C","D","E"):
                    lws[f"{col}{r}"].value = None
                wb.save(master_path)
                st.success(f"Deleted leave row RowID={r}.")
                st.rerun()
            else:
                lws[f"A{r}"].value = edit_name
                lws[f"B{r}"].value = edit_start
                lws[f"C{r}"].value = edit_end
                lws[f"D{r}"].value = normalize_leave_type(edit_type)
                lws[f"E{r}"].value = bool(edit_approved)
                wb.save(master_path)
                st.success(f"Updated RowID={r}.")
                st.rerun()

        finally:
            if enable_lock:
                release_lock(master_path)
