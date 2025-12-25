import streamlit as st
from datetime import date, datetime
from pathlib import Path
import json
import uuid
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Rota Leave Requests (Dropbox-safe)", layout="wide")

st.title("Rota Leave Requests (Dropbox-safe)")
st.write(
    "This UI stores each leave request as an individual JSON file in a Dropbox-synced folder. "
    "This avoids Excel sync conflicts. You can add/edit/delete requests, then compile them into "
    "the `Leave` sheet of `Rota_Publish_Template_ORtools.xlsx` when ready to publish."
)

st.subheader("1) Configure paths (Dropbox)")

default_requests_dir = st.secrets.get("LEAVE_REQUESTS_DIR", "") if hasattr(st, "secrets") else ""
default_workbook_path = st.secrets.get("ROTA_WORKBOOK_PATH", "") if hasattr(st, "secrets") else ""

requests_dir_str = st.text_input(
    "Leave requests folder (Dropbox-synced)",
    value=default_requests_dir,
    placeholder="~/Dropbox/Rota/LeaveRequests",
    help="The app will create one JSON file per request in this folder."
)

workbook_path_str = st.text_input(
    "Rota workbook path (for compile/export)",
    value=default_workbook_path,
    placeholder="~/Dropbox/Rota/Rota_Publish_Template_ORtools.xlsx",
    help="Used only when you press 'Compile into workbook'."
)

requests_dir = Path(requests_dir_str).expanduser() if requests_dir_str else None
workbook_path = Path(workbook_path_str).expanduser() if workbook_path_str else None

col1, col2, col3 = st.columns([1, 1, 2])
with col1:
    create_backup = st.checkbox("Backup workbook on compile", value=True)
with col2:
    replace_leave_sheet = st.checkbox("Replace Leave sheet rows (recommended)", value=True)
with col3:
    st.caption("Recommended: store requests as JSON; only write to Excel during a controlled compile step.")

if not requests_dir_str:
    st.info("Enter a Dropbox folder path for leave requests to continue.")
    st.stop()

requests_dir.mkdir(parents=True, exist_ok=True)

ALLOWED_TYPES = ["Annual", "Study", "NOC"]

def now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

def normalize_leave_type(t: str) -> str:
    t = (t or "").strip()
    mapping = {"annual": "Annual", "study": "Study", "noc": "NOC"}
    return mapping.get(t.lower(), t)

def request_file_path(req_id: str) -> Path:
    return requests_dir / f"{req_id}.json"

def load_requests() -> pd.DataFrame:
    rows = []
    for p in sorted(requests_dir.glob("*.json")):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            rows.append({
                "RequestID": data.get("request_id", p.stem),
                "Name": data.get("name",""),
                "StartDate": pd.to_datetime(data.get("start_date")).date() if data.get("start_date") else None,
                "EndDate": pd.to_datetime(data.get("end_date")).date() if data.get("end_date") else None,
                "LeaveType": data.get("leave_type",""),
                "Approved": bool(data.get("approved", True)),
                "Notes": data.get("notes",""),
                "CreatedAt": data.get("created_at",""),
                "UpdatedAt": data.get("updated_at",""),
                "_path": str(p),
            })
        except Exception:
            continue
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["StartDate","Name"], na_position="last").reset_index(drop=True)
    return df

def validate_dates(s: date, e: date) -> str | None:
    if e < s:
        return "Date to cannot be earlier than Date from."
    return None

def upsert_request(req: dict) -> None:
    request_file_path(req["request_id"]).write_text(json.dumps(req, indent=2), encoding="utf-8")

def delete_request(req_id: str) -> None:
    p = request_file_path(req_id)
    if p.exists():
        p.unlink()

def workbook_backup(path: Path) -> Path | None:
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        bkp = path.with_name(f"{path.stem}_backup_{ts}{path.suffix}")
        bkp.write_bytes(path.read_bytes())
        return bkp
    except Exception:
        return None

def read_consultants_from_workbook(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    if "Consultants" not in wb.sheetnames:
        return []
    ws = wb["Consultants"]
    out = []
    for r in range(2, 2000):
        nm = ws[f"A{r}"].value
        active = ws[f"F{r}"].value
        if nm and bool(active):
            out.append(str(nm))
    return sorted(set(out))

df = load_requests()

consultant_names = []
if workbook_path_str and workbook_path and workbook_path.exists():
    try:
        consultant_names = read_consultants_from_workbook(workbook_path)
    except Exception:
        consultant_names = []

st.subheader("2) Add leave request")
with st.form("add_form"):
    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        if consultant_names:
            name = st.selectbox("Consultant", options=consultant_names)
        else:
            name = st.text_input("Consultant (free text)", value="")
        leave_type = st.selectbox("Leave type", options=ALLOWED_TYPES)
        notes = st.text_input("Notes (optional)", value="")
    with c2:
        start_date = st.date_input("Date from", value=date.today())
    with c3:
        end_date = st.date_input("Date to", value=date.today())
    approved = st.checkbox("Approved", value=True)
    submitted = st.form_submit_button("Create request")

if submitted:
    err = validate_dates(start_date, end_date)
    if err:
        st.error(err)
    elif not name.strip():
        st.error("Consultant name is required.")
    else:
        req_id = str(uuid.uuid4())
        req = {
            "request_id": req_id,
            "name": name.strip(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "leave_type": normalize_leave_type(leave_type),
            "approved": bool(approved),
            "notes": notes.strip(),
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        upsert_request(req)
        st.success(f"Created request {req_id[:8]} for {name}: {start_date} → {end_date} ({leave_type})")
        st.rerun()

st.subheader("3) View / filter requests")

fc1, fc2, fc3, fc4 = st.columns([2, 1, 1, 2])
with fc1:
    filt_name = st.selectbox("Consultant filter", options=["(All)"] + (sorted(df["Name"].unique()) if not df.empty else []))
with fc2:
    filt_type = st.selectbox("Type filter", options=["(All)"] + ALLOWED_TYPES)
with fc3:
    filt_approved = st.selectbox("Approval", options=["(All)", "Approved only", "Not approved"])
with fc4:
    search = st.text_input("Search (name/notes contains)", value="")

view = df.copy()
if not view.empty:
    if filt_name != "(All)":
        view = view[view["Name"] == filt_name]
    if filt_type != "(All)":
        view = view[view["LeaveType"].str.lower() == filt_type.lower()]
    if filt_approved == "Approved only":
        view = view[view["Approved"] == True]
    elif filt_approved == "Not approved":
        view = view[view["Approved"] == False]
    if search.strip():
        s = search.strip().lower()
        view = view[view["Name"].str.lower().str.contains(s) | view["Notes"].str.lower().str.contains(s)]

st.dataframe(view.drop(columns=["_path"]) if not view.empty else view, use_container_width=True, hide_index=True)

st.subheader("4) Edit / delete request")
if df.empty:
    st.info("No requests yet.")
else:
    req_ids = view["RequestID"].tolist() if not view.empty else df["RequestID"].tolist()
    selected = st.selectbox("Select RequestID", options=req_ids)
    row = df[df["RequestID"] == selected].iloc[0].to_dict()
    raw = json.loads(Path(row["_path"]).read_text(encoding="utf-8"))

    with st.form("edit_form"):
        e1, e2, e3 = st.columns([2, 1, 1])
        with e1:
            if consultant_names:
                nm = st.selectbox(
                    "Consultant",
                    options=consultant_names,
                    index=consultant_names.index(raw.get("name","")) if raw.get("name","") in consultant_names else 0
                )
            else:
                nm = st.text_input("Consultant", value=raw.get("name",""))
            lt = st.selectbox(
                "Leave type",
                options=ALLOWED_TYPES,
                index=ALLOWED_TYPES.index(normalize_leave_type(raw.get("leave_type","Annual"))) if normalize_leave_type(raw.get("leave_type","Annual")) in ALLOWED_TYPES else 0
            )
            notes2 = st.text_input("Notes", value=raw.get("notes",""))
        with e2:
            sd = st.date_input("Date from", value=pd.to_datetime(raw.get("start_date")).date() if raw.get("start_date") else date.today(), key="edit_sd")
        with e3:
            ed = st.date_input("Date to", value=pd.to_datetime(raw.get("end_date")).date() if raw.get("end_date") else sd, key="edit_ed")
        appr = st.checkbox("Approved", value=bool(raw.get("approved", True)))

        csave, cdel = st.columns([1, 1])
        with csave:
            save = st.form_submit_button("Save changes")
        with cdel:
            delete = st.form_submit_button("Delete request")

    if save:
        err = validate_dates(sd, ed)
        if err:
            st.error(err)
        elif not nm.strip():
            st.error("Consultant name is required.")
        else:
            raw["name"] = nm.strip()
            raw["leave_type"] = normalize_leave_type(lt)
            raw["start_date"] = sd.isoformat()
            raw["end_date"] = ed.isoformat()
            raw["approved"] = bool(appr)
            raw["notes"] = notes2.strip()
            raw["updated_at"] = now_iso()
            upsert_request(raw)
            st.success("Saved.")
            st.rerun()

    if delete:
        delete_request(selected)
        st.success("Deleted.")
        st.rerun()

st.subheader("5) Compile requests into Excel workbook (Leave sheet)")
st.write(
    "This writes all JSON requests into the workbook's `Leave` sheet. "
    "Run this as a controlled action (typically by the rota administrator)."
)

if not workbook_path_str:
    st.info("Enter the workbook path above to enable compile/export.")
elif workbook_path is None or not workbook_path.exists():
    st.error(f"Workbook not found: {workbook_path}")
else:
    if st.button("Compile into workbook now"):
        try:
            if create_backup:
                bkp = workbook_backup(workbook_path)
                if bkp:
                    st.caption(f"Backup created: {bkp.name}")
                else:
                    st.warning("Backup failed; continuing.")

            wb = load_workbook(workbook_path)
            if "Leave" not in wb.sheetnames:
                st.error("Workbook has no 'Leave' sheet.")
                st.stop()

            lws = wb["Leave"]

            if replace_leave_sheet:
                for r in range(2, 5000):
                    if lws[f"A{r}"].value in (None, ""):
                        break
                    for col in ("A", "B", "C", "D", "E"):
                        lws[f"{col}{r}"].value = None

            df2 = load_requests()
            if df2.empty:
                st.warning("No requests to compile.")
                st.stop()

            r = 2
            for _, rec in df2.iterrows():
                lws[f"A{r}"].value = rec["Name"]
                lws[f"B{r}"].value = rec["StartDate"]
                lws[f"C{r}"].value = rec["EndDate"]
                lws[f"D{r}"].value = normalize_leave_type(rec["LeaveType"])
                lws[f"E{r}"].value = bool(rec["Approved"])
                r += 1

            wb.save(workbook_path)
            st.success(f"Compiled {len(df2)} requests into Leave sheet and saved workbook.")
        except Exception as e:
            st.exception(e)
